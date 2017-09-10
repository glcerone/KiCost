__author__='Giacinto Luigi Cerone'

import future

import sys
import pprint
import re
import difflib
import logging
import tqdm
import os
from random import randint

import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell, xl_range, xl_range_abs

from openpyxl import load_workbook
from openpyxl import Workbook as wbw        #Workbook Writer

class Template:
    
    def __init__(self, templateName):
        
        self.load(templateName)
    
    def load(self, templateName):
        
        self.wbr = load_workbook(filename = templateName, data_only = True)         #Workbook Reader
        self.template = wbr[wbr.get_sheet_names()[0]]
                
    def search(self, sheet, value, pattern='='):
        
        results = dict()
                
        max_row = sheet.max_row
        max_col = sheet.max_column
                
        for row in range(max_row):
            for col in range(max_col):
                actualCell = sheet.cell(row = row+1, column = col+1).value
                
                if actualCell is not None:
                    try:
                        if value in actualCell:
                            results[actualCell.split(pattern)[1]]=(row+1, col+1)
                            print(actualCell.split(pattern)[1])
                    except TypeError:
                        None
        return results
        
    def create_spreadsheet(self, parts, spreadsheet_filename, user_fields, variant, distributors):
        '''Create a spreadsheet using the info for the parts (including their HTML trees).'''
                
        logger.log(DEBUG_OVERVIEW, 'Create spreadsheet...')

        DEFAULT_BUILD_QTY = 100  # Default value for number of boards to build.
        WORKSHEET_NAME = os.path.splitext(os.path.basename(spreadsheet_filename))[0] # Default name for pricing worksheet.
    
        if len(variant) > 0:
            # Append an indication of the variant to the worksheet title.
            # Remove any special characters that might be illegal in a 
            # worksheet name since the variant might be a regular expression.
            WORKSHEET_NAME = WORKSHEET_NAME + '.' + re.sub('[\[\]\\\/\|\?\*\:\(\)]','_',variant)
            
            # Make a list of alphabetically-ordered distributors with web distributors before locals.
            web_dists = sorted([d for d in distributors if distributors[d]['scrape'] != 'local'])
            local_dists = sorted([d for d in distributors if distributors[d]['scrape'] == 'local'])
            dist_list = web_dists + local_dists
    
            # Load the part information from each distributor into the sheet.
            index = 0
            for dist in dist_list:
                dist_start_col = 
                dist_start_row = 
                next_col = add_dist_to_worksheet(wks, wrk_formats, index, START_ROW,
                                                 dist_start_col, TOTAL_COST_ROW,
                                                 refs_col, qty_col, dist, parts)
                index = (index+1) % 2
                # Create a defined range for each set of distributor part data.
                workbook.define_name(
                    '{}_part_data'.format(dist), '={wks_name}!{data_range}'.format(
                        wks_name=WORKSHEET_NAME,
                        data_range=xl_range_abs(START_ROW, dist_start_col,
                                                LAST_PART_ROW, next_col - 1)))

                                    
        results = self.search(template, "Field")

        print(results)
        
        r,c=results['S1-Name']
        
        template.cell(column=c, row=r, value='Ciao')
                
        wbr.save(filename = spreadsheet_filename)

    
        # Create spreadsheet file.
        with xlsxwriter.Workbook(spreadsheet_filename) as workbook:
    
            # Create the various format styles used by various spreadsheet items.
            wrk_formats = {
                'global': workbook.add_format({
                    'font_size': 14,
                    'font_color': 'white',
                    'bold': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'bg_color': '#303030'
                }),
                'digikey': workbook.add_format({
                    'font_size': 14,
                    'font_color': 'white',
                    'bold': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'bg_color': '#CC0000'  # Digi-Key red.
                }),
                'mouser': workbook.add_format({
                    'font_size': 14,
                    'font_color': 'white',
                    'bold': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'bg_color': '#004A85'  # Mouser blue.
                }),
                'newark': workbook.add_format({
                    'font_size': 14,
                    'font_color': 'white',
                    'bold': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'bg_color': '#A2AE06'  # Newark/E14 olive green.
                }),
                'rs': workbook.add_format({
                    'font_size': 14,
                    'font_color': 'white',
                    'bold': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'bg_color': '#FF0000'  # RS Components red.
                }),
                'farnell': workbook.add_format({
                    'font_size': 14,
                    'font_color': 'white',
                    'bold': True,
                    'align': 'center',
                    'valign': 'vcenter',
                    'bg_color': '#FF6600'  # Farnell/E14 orange.
                }),
    
                'local_lbl': [
                    workbook.add_format({
                        'font_size': 14,
                        'font_color': 'black',
                        'bold': True,
                        'align': 'center',
                        'valign': 'vcenter',
                        'bg_color': '#909090'  # Darker grey.
                    }),
                    workbook.add_format({
                        'font_size': 14,
                        'font_color': 'black',
                        'bold': True,
                        'align': 'center',
                        'valign': 'vcenter',
                        'bg_color': '#c0c0c0'  # Lighter grey.
                    }),
                ],
                'header': workbook.add_format({
                    'font_size': 12,
                    'bold': True,
                    'align': 'center',
                    'valign': 'top',
                    'text_wrap': True
                }),
                'board_qty': workbook.add_format(
                    {'font_size': 13,
                     'bold': True,
                     'align': 'right'}),
                'total_cost_label': workbook.add_format({
                    'font_size': 13,
                    'bold': True,
                    'align': 'right',
                    'valign': 'vcenter'}),
                'unit_cost_label': workbook.add_format({
                    'font_size': 13,
                    'bold': True,
                    'align': 'right',
                    'valign': 'vcenter'
                }),
                'total_cost_currency': workbook.add_format({
                    'font_size': 13,
                    'font_color': 'red',
                    'bold': True,
                    'num_format': '$#,##0.00',
                    'valign': 'vcenter'}),
                'unit_cost_currency': workbook.add_format({
                    'font_size': 13,
                    'font_color': 'green',
                    'bold': True,
                    'num_format': '$#,##0.00',
                    'valign': 'vcenter'
                }),
                'best_price': workbook.add_format({'bg_color': '#80FF80', }),
                'currency': workbook.add_format({'num_format': '$#,##0.00'}),
                'centered_text': workbook.add_format({'align': 'center'}),
            }
    
            # Create the worksheet that holds the pricing information.
            wks = workbook.add_worksheet(WORKSHEET_NAME)
    
            # Set the row & column for entering the part information in the sheet.
            START_COL = 0
            BOARD_QTY_ROW = 0
            TOTAL_COST_ROW = BOARD_QTY_ROW + 1
            UNIT_COST_ROW = TOTAL_COST_ROW + 1
            START_ROW = 4
            LABEL_ROW = START_ROW + 1
            COL_HDR_ROW = LABEL_ROW + 1
            FIRST_PART_ROW = COL_HDR_ROW + 1
            LAST_PART_ROW = COL_HDR_ROW + len(parts) - 1
    
            # Load the global part information (not distributor-specific) into the sheet.
            # next_col = the column immediately to the right of the global data.
            # qty_col = the column where the quantity needed of each part is stored.
            next_col, refs_col, qty_col = add_globals_to_worksheet(
                wks, wrk_formats, START_ROW, START_COL, TOTAL_COST_ROW, parts, user_fields)
            # Create a defined range for the global data.
            workbook.define_name(
                'global_part_data', '={wks_name}!{data_range}'.format(
                    wks_name=WORKSHEET_NAME,
                    data_range=xl_range_abs(START_ROW, START_COL, LAST_PART_ROW,
                                            next_col - 1)))
    
            # Create the cell where the quantity of boards to assemble is entered.
            # Place the board qty cells near the right side of the global info.
            wks.write(BOARD_QTY_ROW, next_col - 2, 'Board Qty:',
                      wrk_formats['board_qty'])
            wks.write(BOARD_QTY_ROW, next_col - 1, DEFAULT_BUILD_QTY,
                      wrk_formats['board_qty'])  # Set initial board quantity.
            # Define the named cell where the total board quantity can be found.
            workbook.define_name('BoardQty', '={wks_name}!{cell_ref}'.format(
                wks_name=WORKSHEET_NAME,
                cell_ref=xl_rowcol_to_cell(BOARD_QTY_ROW, next_col - 1,
                                           row_abs=True,
                                           col_abs=True)))
    
            # Create the row to show total cost of board parts for each distributor.
            wks.write(TOTAL_COST_ROW, next_col - 2, 'Total Cost:',
                      wrk_formats['total_cost_label'])
    
            # Define the named cell where the total cost can be found.
            workbook.define_name('TotalCost', '={wks_name}!{cell_ref}'.format(
                wks_name=WORKSHEET_NAME,
                cell_ref=xl_rowcol_to_cell(TOTAL_COST_ROW, next_col - 1,
                                           row_abs=True,
                                           col_abs=True)))
    
    
            # Create the row to show unit cost of board parts.
            wks.write(UNIT_COST_ROW, next_col - 2, 'Unit Cost:',
                      wrk_formats['unit_cost_label'])
            wks.write(UNIT_COST_ROW, next_col - 1, "=TotalCost/BoardQty",
                      wrk_formats['unit_cost_currency'])
    
            # Freeze view of the global information and the column headers, but
            # allow the distributor-specific part info to scroll.
            wks.freeze_panes(COL_HDR_ROW, next_col)
    
            # Make a list of alphabetically-ordered distributors with web distributors before locals.
            web_dists = sorted([d for d in distributors if distributors[d]['scrape'] != 'local'])
            local_dists = sorted([d for d in distributors if distributors[d]['scrape'] == 'local'])
            dist_list = web_dists + local_dists
    
            # Load the part information from each distributor into the sheet.
            index = 0
            for dist in dist_list:
                dist_start_col = next_col
                next_col = add_dist_to_worksheet(wks, wrk_formats, index, START_ROW,
                                                 dist_start_col, TOTAL_COST_ROW,
                                                 refs_col, qty_col, dist, parts)
                index = (index+1) % 2
                # Create a defined range for each set of distributor part data.
                workbook.define_name(
                    '{}_part_data'.format(dist), '={wks_name}!{data_range}'.format(
                        wks_name=WORKSHEET_NAME,
                        data_range=xl_range_abs(START_ROW, dist_start_col,
                                                LAST_PART_ROW, next_col - 1)))
    def add_dist_to_worksheet(wks, wrk_formats, index, start_row, start_col,
                              total_cost_row, part_ref_col, part_qty_col, dist,
                              parts):
        '''Add distributor-specific part data to the spreadsheet.'''
    
        # Columns for the various types of distributor-specific part data.
        columns = {
            'avail': {
                'col': 0,
                # column offset within this distributor range of the worksheet.
                'level': 1,  # Outline level (or hierarchy level) for this column.
                'label': 'Avail',  # Column header label.
                'width': None,  # Column width (default in this case).
                'comment': 'Available quantity of each part at the distributor.'
                # Column header tool-tip.
            },
            'purch': {
                'col': 1,
                'level': 2,
                'label': 'Purch',
                'width': None,
                'comment': 'Purchase quantity of each part from this distributor.'
            },
            'unit_price': {
                'col': 2,
                'level': 2,
                'label': 'Unit$',
                'width': None,
                'comment': 'Unit price of each part from this distributor.'
            },
            'ext_price': {
                'col': 3,
                'level': 0,
                'label': 'Ext$',
                'width': 15,  # Displays up to $9,999,999.99 without "###".
                'comment':
                '(Unit Price) x (Purchase Qty) of each part from this distributor.'
            },
            'part_num': {
                'col': 4,
                'level': 2,
                'label': 'Cat#',
                'width': None,
                'comment': 'Distributor-assigned part number for each part.'
            },
            'part_url': {
                'col': 5,
                'level': 2,
                'label': 'Doc',
                'width': None,
                'comment': 'Link to distributor webpage for each part.'
            },
        }
        num_cols = len(list(columns.keys()))
    
        row = start_row  # Start building distributor section at this row.
    
        # Add label for this distributor.
        try:
            wks.merge_range(row, start_col, row, start_col + num_cols - 1,
                        distributors[dist]['label'].title(), wrk_formats[dist])
        except KeyError:
            wks.merge_range(row, start_col, row, start_col + num_cols - 1,
                        distributors[dist]['label'].title(), wrk_formats['local_lbl'][index])
        row += 1  # Go to next row.
    
        # Add column headers, comments, and outline level (for hierarchy).
        for k in list(columns.keys()):
            col = start_col + columns[k]['col']  # Column index for this column.
            wks.write_string(row, col, columns[k]['label'], wrk_formats['header'])
            wks.write_comment(row, col, columns[k]['comment'])
            wks.set_column(col, col, columns[k]['width'], None,
                           {'level': columns[k]['level']})
        row += 1  # Go to next row.
    
        num_parts = len(parts)
    
        # Add distributor data for each part.
        PART_INFO_FIRST_ROW = row  # Starting row of part info.
        PART_INFO_LAST_ROW = PART_INFO_FIRST_ROW + num_parts - 1  # Last row of part info.
    
        for part in parts:
    
            # Get the distributor part number.
            dist_part_num = part.part_num[dist]
    
            # Extract price tiers from distributor HTML page tree.
            price_tiers = part.price_tiers[dist]
    
            # Enter a link to the distributor webpage for this part, even if there
            # is no valid quantity or pricing for the part (see next conditional).
            # Having the link present will help debug if the extraction of the
            # quantity or pricing information was done correctly.
            if part.url[dist]:
                wks.write_url(row, start_col + columns['part_url']['col'],
                          part.url[dist], wrk_formats['centered_text'],
                          string='Link')
    
            # If the part number doesn't exist or the part quantity is None 
            # (not the same as 0), then the distributor doesn't stock this part
            # so leave this row blank.
            # Also leave this row blank if the pricing info doesn't exist so a $0
            # price isn't injected that messes up the search for the best price.
            if len(dist_part_num) == 0 or part.qty_avail[dist] is None or len(list(price_tiers.keys())) == 0:
                row += 1  # Skip this row and go to the next.
                continue
    
            # Enter distributor part number for ordering purposes.
            wks.write(row, start_col + columns['part_num']['col'], dist_part_num,
                      None)
    
            # Enter quantity of part available at this distributor.
            wks.write(row, start_col + columns['avail']['col'],
                      part.qty_avail[dist], None)
    
            # Purchase quantity always starts as blank because nothing has been purchased yet.
            wks.write(row, start_col + columns['purch']['col'], '', None)
    
            # Add the price for a single unit if it doesn't already exist in the tiers.
            try:
                min_qty = min(price_tiers.keys())
                if min_qty > 1:
                    price_tiers[1] = price_tiers[
                        min_qty
                    ]  # Set unit price to price of lowest available quantity.
            except ValueError:  # This happens if the price tier list is empty.
                pass
            price_tiers[0] = 0.00  # Enter quantity-zero pricing so LOOKUP works correctly in the spreadsheet.
    
            # Sort the tiers based on quantities and turn them into lists of strings.
            qtys = sorted(price_tiers.keys())
            prices = [str(price_tiers[q]) for q in qtys]
            qtys = [str(q) for q in qtys]
    
            purch_qty_col = start_col + columns['purch']['col']
            unit_price_col = start_col + columns['unit_price']['col']
            ext_price_col = start_col + columns['ext_price']['col']
    
            # Enter a spreadsheet lookup function that determines the unit price based on the needed quantity
            # or the purchased quantity (if that is non-zero).
            wks.write_formula(
                row, unit_price_col,
                '=iferror(lookup(if({purch_qty}="",{needed_qty},{purch_qty}),{{{qtys}}},{{{prices}}}),"")'.format(
                    needed_qty=xl_rowcol_to_cell(row, part_qty_col),
                    purch_qty=xl_rowcol_to_cell(row, purch_qty_col),
                    qtys=','.join(qtys),
                    prices=','.join(prices)), wrk_formats['currency'])
            # Conditionally format the unit price cell that contains the best price.
            wks.conditional_format(row, unit_price_col, row, unit_price_col, {
                'type': 'cell',
                'criteria': '<=',
                'value': xl_rowcol_to_cell(row, 7),
                # This is the global data cell holding the minimum unit price for this part.
                'format': wrk_formats['best_price']
            })
    
            # Enter the formula for the extended price = purch qty * unit price.
            wks.write_formula(
                row, ext_price_col,
                '=iferror(if({purch_qty}="",{needed_qty},{purch_qty})*{unit_price},"")'.format(
                    needed_qty=xl_rowcol_to_cell(row, part_qty_col),
                    purch_qty=xl_rowcol_to_cell(row, purch_qty_col),
                    unit_price=xl_rowcol_to_cell(row, unit_price_col)),
                wrk_formats['currency'])
            # Conditionally format the extended price cell that contains the best price.
            wks.conditional_format(row, ext_price_col, row, ext_price_col, {
                'type': 'cell',
                'criteria': '<=',
                'value': xl_rowcol_to_cell(row, 8),
                # This is the global data cell holding the minimum extended price for this part.
                'format': wrk_formats['best_price']
            })
    
            # Finished processing distributor data for this part.
            row += 1  # Go to next row.
    
        # Sum the extended prices for all the parts to get the total cost from this distributor.
        total_cost_col = start_col + columns['ext_price']['col']
        wks.write(total_cost_row, total_cost_col, '=sum({sum_range})'.format(
            sum_range=xl_range(PART_INFO_FIRST_ROW, total_cost_col,
                               PART_INFO_LAST_ROW, total_cost_col)),
                  wrk_formats['total_cost_currency'])
    
        # Add list of part numbers and purchase quantities for ordering from this distributor.
        ORDER_START_COL = start_col + 1
        ORDER_FIRST_ROW = PART_INFO_LAST_ROW + 2
        ORDER_LAST_ROW = ORDER_FIRST_ROW + num_parts - 1
    
        # Each distributor has a different format for entering ordering information,
        # so we account for that here.
        order_col = {}
        order_col_numeric = {}
        order_delimiter = {}
        dist_col = {}
        for position, col_tag in enumerate(distributors[dist]['order_cols']):
            order_col[col_tag] = ORDER_START_COL + position  # Column for this order info.
            order_col_numeric[col_tag] = (col_tag ==
                                          'purch')  # Is this order info numeric?
            order_delimiter[col_tag] = distributors[dist][
                'order_delimiter'
            ]  # Delimiter btwn order columns.
            # For the last column of order info, the delimiter is blanked.
            if position + 1 == len(distributors[dist]['order_cols']):
                order_delimiter[col_tag] = ''
            # If the column tag doesn't exist in the list of distributor columns,
            # then assume its for the part reference column in the global data section
            # of the worksheet.
            try:
                dist_col[col_tag] = start_col + columns[col_tag]['col']
            except KeyError:
                dist_col[col_tag] = part_ref_col
    
        def enter_order_info(info_col, order_col, numeric=False, delimiter=''):
            ''' This function enters a function into a spreadsheet cell that
                prints the information found in info_col into the order_col column
                of the order.
            '''
    
            # This very complicated spreadsheet function does the following:
            # 1) Computes the set of row indices in the part data that have
            #    non-empty cells in sel_range1 and sel_range2. (Innermost
            #    nested IF and ROW commands.) sel_range1 and sel_range2 are
            #    the part's catalog number and purchase quantity.
            # 2) Selects the k'th smallest of the row indices where k is the
            #    number of rows between the current part row in the order and the
            #    top row of the order. (SMALL() and ROW() commands.)
            # 3) Gets the cell contents  from the get_range using the k'th
            #    smallest row index found in step #2. (INDEX() command.)
            # 4) Converts the cell contents to a string if it is numeric.
            #    (num_to_text_func is used.) Otherwise, it's already a string.
            # 5) CONCATENATES the string from step #4 with the delimiter
            #    that goes between fields of an order for a part.
            #    (CONCATENATE() command.)
            # 6) If any error occurs (which usually means the indexed cell
            #    contents were blank), then a blank is printed. Otherwise,
            #    the string from step #5 is printed in this cell.
            order_info_func = '''
                IFERROR(
                    CONCATENATE(
                        {num_to_text_func}(
                            INDEX(
                                {get_range},
                                SMALL(
                                    IF(
                                        {sel_range2} <> "",
                                        IF(
                                            {sel_range1} <> "",
                                            ROW({sel_range1}) - MIN(ROW({sel_range1})) + 1,
                                            ""
                                        ),
                                        ""
                                    ),
                                    ROW()-ROW({order_first_row})+1
                                )
                            )
                            {num_to_text_fmt}
                        ),
                        {delimiter}
                    ),
                    ""
                )
            '''
    
            # Strip all the whitespace from the function string.
            order_info_func = re.sub('[\s\n]', '', order_info_func)
    
            # This sets the function and conversion format to use if
            # numeric cell contents have to be converted to a string.
            if numeric:
                num_to_text_func = 'TEXT'
                num_to_text_fmt = ',"##0"'
            else:
                num_to_text_func = ''
                num_to_text_fmt = ''
    
            # This puts the order column delimiter into a form acceptable in a spreadsheet formula.
            if delimiter != '':
                delimiter = '"{}"'.format(delimiter)
    
            # These are the columns where the part catalog numbers and purchase quantities can be found.
            purch_qty_col = start_col + columns['purch']['col']
            part_num_col = start_col + columns['part_num']['col']
    
            # Now write the order_info_func into every row of the order in the given column.
            for r in range(ORDER_FIRST_ROW, ORDER_LAST_ROW + 1):
                wks.write_array_formula(
                    xl_range(r, order_col, r, order_col),
                    '{{={func}}}'.format(func=order_info_func.format(
                        order_first_row=xl_rowcol_to_cell(ORDER_FIRST_ROW, 0,
                                                          row_abs=True),
                        sel_range1=xl_range_abs(PART_INFO_FIRST_ROW, purch_qty_col,
                                                PART_INFO_LAST_ROW, purch_qty_col),
                        sel_range2=xl_range_abs(PART_INFO_FIRST_ROW, part_num_col,
                                                PART_INFO_LAST_ROW, part_num_col),
                        get_range=xl_range_abs(PART_INFO_FIRST_ROW, info_col,
                                               PART_INFO_LAST_ROW, info_col),
                        delimiter=delimiter,
                        num_to_text_func=num_to_text_func,
                        num_to_text_fmt=num_to_text_fmt)))
    
        # For every column in the order info range, enter the part order information.
        for col_tag in ('purch', 'part_num', 'refs'):
            enter_order_info(dist_col[col_tag], order_col[col_tag],
                             numeric=order_col_numeric[col_tag],
                             delimiter=order_delimiter[col_tag])
    
        return start_col + num_cols  # Return column following the globals so we know where to start next set of cells.



if __name__ == '__main__':
    t=Template('template2.xlsx')
    #~ t.read()
    #~ t.create_spreadsheet()
